"""Shared Excel/CSV export helpers using openpyxl."""
from datetime import datetime
from decimal import Decimal
import csv

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill(start_color='C46616', end_color='C46616',
                          fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _safe_cell(value):
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, 'amount'):
        return float(value.amount)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def excel_response(filename, headers, rows, sheet_name='Report'):
    """rows is iterable of lists matching headers order."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append([_safe_cell(v) for v in row])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def csv_response(filename, headers, rows):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(resp)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_safe_cell(v) for v in row])
    return resp
